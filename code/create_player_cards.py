import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import re
import os

def extract_year_and_team(a1_value):
    """Extract year and team name from A1 cell value"""
    if pd.isna(a1_value):
        return None, None
    
    # Convert to string and clean up
    text = str(a1_value).strip()
    
    # Look for year pattern (4 digits)
    year_match = re.search(r'\b(19|20)\d{2}\b', text)
    year = year_match.group() if year_match else None
    
    # Extract team name (everything after the year)
    if year:
        team_part = text.split(year, 1)[1].strip()
        # Clean up any extra characters
        team = re.sub(r'[^\w\s]', '', team_part).strip()
    else:
        team = None
    
    return year, team

def get_card_position(card_index, is_batter=True):
    """Calculate the position of a card on the page (3 cards per row)"""
    cards_per_row = 3
    
    # Calculate row and column position
    row = card_index // cards_per_row
    col = card_index % cards_per_row
    
    # Each card is 8 rows tall, so multiply by 8
    start_row = row * 8 + 1
    start_col = col * 7 + 1
    
    return start_row, start_col

def copy_template_formatting(ws, template_ws, start_row, start_col, is_batter=True):
    """Copy formatting from template to the specified position"""
    # Copy the 8x7 card area formatting
    for row in range(8):
        for col in range(7):
            source_cell = template_ws.cell(row=row+1, column=col+1)
            target_cell = ws.cell(row=start_row+row, column=start_col+col)
            
            # Copy formatting
            if hasattr(source_cell, 'font') and source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            if hasattr(source_cell, 'alignment') and source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # Copy border properties individually
            if hasattr(source_cell, 'border') and source_cell.border:
                try:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom
                    )
                except:
                    # If border copying fails, use default thin border
                    target_cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Copy fill properties individually
            if hasattr(source_cell, 'fill') and source_cell.fill:
                try:
                    from openpyxl.styles import PatternFill
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
                except:
                    pass  # Skip fill if copying fails
            
            if hasattr(source_cell, 'number_format') and source_cell.number_format:
                target_cell.number_format = source_cell.number_format
    
    # Copy merged ranges from template
    for merged_range in template_ws.merged_cells.ranges:
        # Calculate the offset for this card
        new_start_row = start_row + merged_range.min_row - 1
        new_end_row = start_row + merged_range.max_row - 1
        new_start_col = start_col + merged_range.min_col - 1
        new_end_col = start_col + merged_range.max_col - 1
        
        # Apply the same merge to the target worksheet
        ws.merge_cells(
            start_row=new_start_row,
            start_column=new_start_col,
            end_row=new_end_row,
            end_column=new_end_col
        )

def create_card_structure(ws, template_ws, start_row, start_col, is_batter=True):
    """Create the basic card structure with template formatting"""
    # Copy formatting from template
    copy_template_formatting(ws, template_ws, start_row, start_col, is_batter)
    
    # Add fixed text labels (these will override template values but keep formatting)
    # Age: A3
    ws.cell(row=start_row+2, column=start_col, value="Age:")
    
    # Positions: D3
    ws.cell(row=start_row+2, column=start_col+3, value="Positions:")
    
    # Bats: A4
    ws.cell(row=start_row+3, column=start_col, value="Bats:")
    
    # Throws: D4
    ws.cell(row=start_row+3, column=start_col+3, value="Throws:")
    
    # DEFENSE: A5
    ws.cell(row=start_row+4, column=start_col, value="DEFENSE:")
    
    if is_batter:
        # OFFENSE: A6 (for batters)
        ws.cell(row=start_row+5, column=start_col, value="OFFENSE:")
        
        # PROBABLE HIT: A7 (for batters)
        ws.cell(row=start_row+6, column=start_col, value="PROBABLE HIT:")
    else:
        # PITCHING: A6 (for pitchers)
        ws.cell(row=start_row+5, column=start_col, value="PITCHING:")
        
        # CONTROL #: A7 (for pitchers)
        ws.cell(row=start_row+6, column=start_col, value="CONTROL #:")
        
        # PROBABLE HIT: A8 (for pitchers)
        ws.cell(row=start_row+7, column=start_col, value="PROBABLE HIT:")

def populate_batter_card(ws, start_row, start_col, player_data, year, team):
    """Populate a batter card with player data"""
    # Year in G8 (7th column, 8th row)
    cell = ws.cell(row=start_row+7, column=start_col+6)
    if not isinstance(cell, MergedCell):
        cell.value = year
    
    # Team name in A1 (1st row, 1st column) - merge A1:G1
    cell = ws.cell(row=start_row, column=start_col)
    if not isinstance(cell, MergedCell):
        cell.value = team
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+6)
    
    # Player name in A2 (2nd row, 1st column) - merge A2:E2
    cell = ws.cell(row=start_row+1, column=start_col)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Player']
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+4)
    
    # Primary position in F2:G2 (2nd row, 6th-7th column) - merge F2:G2
    cell = ws.cell(row=start_row+1, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Primary']
    ws.merge_cells(start_row=start_row+1, start_column=start_col+5, end_row=start_row+1, end_column=start_col+6)
    
    # Age in C3 (3rd row, 3rd column)
    cell = ws.cell(row=start_row+2, column=start_col+2)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Age']
    
    # Age label in A3:B3 - merge A3:B3 and right justify
    age_cell = ws.cell(row=start_row+2, column=start_col)
    age_cell.value = "Age:"
    age_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)
    
    # Positions in F3 (3rd row, 6th column) - merge F3:G3
    cell = ws.cell(row=start_row+2, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Positions']
    ws.merge_cells(start_row=start_row+2, start_column=start_col+5, end_row=start_row+2, end_column=start_col+6)
    
    # Bats in C4 (4th row, 3rd column)
    cell = ws.cell(row=start_row+3, column=start_col+2)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['B']
    
    # Bats label in A4:B4 - merge A4:B4 and right justify
    bats_cell = ws.cell(row=start_row+3, column=start_col)
    bats_cell.value = "Bats:"
    bats_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+3, start_column=start_col, end_row=start_row+3, end_column=start_col+1)
    
    # Throws in F4 (4th row, 6th column)
    cell = ws.cell(row=start_row+3, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['T']
    
    # Defensive rating in D5 (5th row, 4th column) - merge D5:G5
    cell = ws.cell(row=start_row+4, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['DEF']
    ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+6)
    
    # DEFENSE label in A5:B5 - merge A5:B5 and right justify
    def_cell = ws.cell(row=start_row+4, column=start_col)
    def_cell.value = "DEFENSE:"
    def_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+4, start_column=start_col, end_row=start_row+4, end_column=start_col+1)
    
    # Offensive rating in D6 (6th row, 4th column) - merge D6:G6
    cell = ws.cell(row=start_row+5, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Batter Rating']
    ws.merge_cells(start_row=start_row+5, start_column=start_col+3, end_row=start_row+5, end_column=start_col+6)
    
    # Probable hit number in D7 (7th row, 4th column)
    cell = ws.cell(row=start_row+6, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['BPH']

def populate_pitcher_card(ws, start_row, start_col, player_data, year, team):
    """Populate a pitcher card with player data"""
    # Year in G8 (7th column, 8th row)
    cell = ws.cell(row=start_row+7, column=start_col+6)
    if not isinstance(cell, MergedCell):
        cell.value = year
    
    # Team name in A1 (1st row, 1st column) - merge A1:G1
    cell = ws.cell(row=start_row, column=start_col)
    if not isinstance(cell, MergedCell):
        cell.value = team
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+6)
    
    # Player name in A2 (2nd row, 1st column) - merge A2:E2
    cell = ws.cell(row=start_row+1, column=start_col)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Player']
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+4)
    
    # Primary position in F2:G2 (2nd row, 6th-7th column) - merge F2:G2
    cell = ws.cell(row=start_row+1, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Primary']
    ws.merge_cells(start_row=start_row+1, start_column=start_col+5, end_row=start_row+1, end_column=start_col+6)
    
    # Age in C3 (3rd row, 3rd column)
    cell = ws.cell(row=start_row+2, column=start_col+2)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Age']
    
    # Age label in A3:B3 - merge A3:B3 and right justify
    age_cell = ws.cell(row=start_row+2, column=start_col)
    age_cell.value = "Age:"
    age_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+1)
    
    # Positions in F3 (3rd row, 6th column) - merge F3:G3
    cell = ws.cell(row=start_row+2, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Positions']
    ws.merge_cells(start_row=start_row+2, start_column=start_col+5, end_row=start_row+2, end_column=start_col+6)
    
    # Bats in C4 (4th row, 3rd column)
    cell = ws.cell(row=start_row+3, column=start_col+2)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['B']
    
    # Bats label in A4:B4 - merge A4:B4 and right justify
    bats_cell = ws.cell(row=start_row+3, column=start_col)
    bats_cell.value = "Bats:"
    bats_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+3, start_column=start_col, end_row=start_row+3, end_column=start_col+1)
    
    # Throws in F4 (4th row, 6th column)
    cell = ws.cell(row=start_row+3, column=start_col+5)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['T']
    
    # Defensive rating in D5 (5th row, 4th column) - merge D5:G5 and left justify
    cell = ws.cell(row=start_row+4, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['DEF']
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+6)
    
    # DEFENSE label in A5:B5 - merge A5:B5 and right justify
    def_cell = ws.cell(row=start_row+4, column=start_col)
    def_cell.value = "DEFENSE:"
    def_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=start_row+4, start_column=start_col, end_row=start_row+4, end_column=start_col+1)
    
    # Pitcher rating in D6 (6th row, 4th column) - merge D6:G6 and left justify
    cell = ws.cell(row=start_row+5, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['Pitcher Rating']
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=start_row+5, start_column=start_col+3, end_row=start_row+5, end_column=start_col+6)
    
    # Control number in D7 (7th row, 4th column)
    cell = ws.cell(row=start_row+6, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['PCN']
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Probable hit number in D8 (8th row, 4th column)
    cell = ws.cell(row=start_row+7, column=start_col+3)
    if not isinstance(cell, MergedCell):
        cell.value = player_data['PPH']
        cell.alignment = Alignment(horizontal='left', vertical='center')

def create_player_cards(data_file, template_file, output_file):
    """Main function to create player cards"""
    # Load template
    try:
        template_wb = openpyxl.load_workbook(template_file)
        batter_template = template_wb['Batter Card Template']
        pitcher_template = template_wb['Pitcher Card Template']
        print(f"Successfully loaded templates: {template_file}")
    except Exception as e:
        print(f"Error loading template: {e}")
        return
    
    # Load data file
    data_wb = openpyxl.load_workbook(data_file)
    
    # Create output workbook
    output_wb = Workbook()
    
    # Process each team
    for sheet_name in data_wb.sheetnames:
        if sheet_name == 'TOT':  # Skip total sheet
            continue
            
        print(f"Processing team: {sheet_name}")
        
        # Read team data
        team_sheet = data_wb[sheet_name]
        
        # Get year and team name from A1
        a1_value = team_sheet['A1'].value
        year, team = extract_year_and_team(a1_value)
        
        if not year or not team:
            print(f"Warning: Could not extract year/team from A1 for {sheet_name}")
            continue
        
        # Create new sheet for this team
        team_sheet_output = output_wb.create_sheet(title=sheet_name)
        
        # Turn off gridlines
        team_sheet_output.sheet_view.showGridLines = False
        
        # Set all column widths to 29 pixels using the correct conversion
        desired_pixel_width = 29
        excel_width_unit = desired_pixel_width / 6.1
        for col in range(1, 50):  # Set width for first 50 columns
            team_sheet_output.column_dimensions[get_column_letter(col)].width = excel_width_unit
        
        # Read data into DataFrame for easier processing
        data = []
        for row in range(3, team_sheet.max_row + 1):  # Skip header row (row 2)
            row_data = {}
            for col in range(1, team_sheet.max_column + 1):
                cell_value = team_sheet.cell(row=row, column=col).value
                row_data[get_column_letter(col)] = cell_value
            data.append(row_data)
        
        df = pd.DataFrame(data)
        
        # Process batters (non-blank Batter Rating column J)
        batters = df[df['J'].notna() & (df['J'] != '')]
        print(f"  Found {len(batters)} batters")
        
        # Process pitchers (non-blank Pitcher Rating column M)
        pitchers = df[df['M'].notna() & (df['M'] != '')]
        print(f"  Found {len(pitchers)} pitchers")
        
        # Separate only-batters from pitcher-batters
        only_batters = batters[~batters.index.isin(pitchers.index)]
        pitcher_batters = batters[batters.index.isin(pitchers.index)]
        
        print(f"  Found {len(only_batters)} only-batters")
        print(f"  Found {len(pitcher_batters)} pitcher-batters")
        print(f"  Found {len(pitchers)} pitchers")
        
        # Create cards in the specified order
        card_index = 0
        
        # 1. Only-batters first
        print(f"  Creating {len(only_batters)} only-batter cards...")
        for i, batter in only_batters.iterrows():
            start_row, start_col = get_card_position(card_index, is_batter=True)
            
            # Create card structure with template formatting
            create_card_structure(team_sheet_output, batter_template, start_row, start_col, is_batter=True)
            
            # Populate with data
            player_data = {
                'Player': batter['B'],
                'Primary': batter['Q'],
                'Age': batter['C'],
                'Positions': batter['H'],
                'B': batter['L'],
                'T': batter['P'],
                'DEF': batter['I'],
                'Batter Rating': batter['J'],
                'BPH': batter['K']
            }
            
            populate_batter_card(team_sheet_output, start_row, start_col, player_data, year, team)
            card_index += 1
        
        # 2. Pitcher-batters (batter cards for players who also pitch)
        print(f"  Creating {len(pitcher_batters)} pitcher-batter cards...")
        for i, batter in pitcher_batters.iterrows():
            start_row, start_col = get_card_position(card_index, is_batter=True)
            
            # Create card structure with template formatting
            create_card_structure(team_sheet_output, batter_template, start_row, start_col, is_batter=True)
            
            # Populate with data
            player_data = {
                'Player': batter['B'],
                'Primary': batter['Q'],
                'Age': batter['C'],
                'Positions': batter['H'],
                'B': batter['L'],
                'T': batter['P'],
                'DEF': batter['I'],
                'Batter Rating': batter['J'],
                'BPH': batter['K']
            }
            
            populate_batter_card(team_sheet_output, start_row, start_col, player_data, year, team)
            card_index += 1
        
        # 3. Pitcher cards (for players with pitcher ratings)
        print(f"  Creating {len(pitchers)} pitcher cards...")
        for i, pitcher in pitchers.iterrows():
            start_row, start_col = get_card_position(card_index, is_batter=False)
            
            # Create card structure with template formatting
            create_card_structure(team_sheet_output, pitcher_template, start_row, start_col, is_batter=False)
            
            # Populate with data
            player_data = {
                'Player': pitcher['B'],
                'Primary': pitcher['Q'],
                'Age': pitcher['C'],
                'Positions': pitcher['H'],
                'B': pitcher['L'],
                'T': pitcher['P'],
                'DEF': pitcher['I'],
                'Pitcher Rating': pitcher['M'],
                'PCN': pitcher['N'],
                'PPH': pitcher['O']
            }
            
            populate_pitcher_card(team_sheet_output, start_row, start_col, player_data, year, team)
            card_index += 1
    
    # Remove default sheet
    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])
    
    # Save output file
    output_wb.save(output_file)
    print(f"Player cards saved to: {output_file}")

if __name__ == "__main__":
    # File paths
    data_file = "../data/1995 rosters updated formatted v3.xlsx"
    template_file = "../data/Player Cards Template.xlsx"
    output_file = "../data/1995 Player Cards v3.xlsx"
    
    # Check if files exist
    if not os.path.exists(data_file):
        print(f"Error: Data file {data_file} not found!")
        exit(1)
    
    if not os.path.exists(template_file):
        print(f"Error: Template file {template_file} not found!")
        exit(1)
    
    # Create player cards
    create_player_cards(data_file, template_file, output_file) 